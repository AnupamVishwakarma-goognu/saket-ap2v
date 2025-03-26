import openpyxl
import threading as th
import logging

class Excel_Features(object):
    def load_excel_file(**kwargs):
        data = {}
        read_only_value = kwargs['read_only']
        for key, value in kwargs.items():
            if key != "read_only":
                if key == "Followups_excel_sheet":
                    data['followup_load_excel'] = openpyxl.load_workbook(value, read_only=read_only_value)
                elif key == "Enquiry_excel_sheet":
                    data['enquiry_load_excel'] = openpyxl.load_workbook(value, read_only=read_only_value)
        return data

    def active_excel_load_data(excel_file_load_data_dict=None):
        data = {}
        if excel_file_load_data_dict != None:
            for key, value in excel_file_load_data_dict.items():
                data[key] = value.active
        return data

    def read_excel_data(excel_data=None):
        if excel_data != None:
            for key, value in excel_data.items():
                if key == "followup_load_excel":
                    Excel_Features.followups_excel_max_row = excel_data[key].max_row
                elif key == "enquiry_load_excel":
                    Excel_Features.enquiry_excel_max_row = excel_data[key].max_row

    def read_email_data_using_thread(data=None):
        key, key_value = data
        if key == "followup_load_excel":
            DataRead.data[key] = [i[5] for i in key_value.values if i[5] != 'Enquiry Email']
        elif key == "enquiry_load_excel":
            DataRead.data[key] = [i[4] for i in key_value.values if i[4] != "Email"]

    def parse_email_data_to_thread(excel_active_data=None):
        for excel_active_tuple_data in excel_active_data.items():
            thread_for_read_email_data = th.Thread(target=Excel_Features.read_email_data_using_thread, args=(excel_active_tuple_data,))
            thread_for_read_email_data.start()
            thread_for_read_email_data.join()

class DataRead:
    data = {}
    def __init__(self, followup_excel=None, enquiry_excel=None, read_only=True):
        self.followup_excel = followup_excel
        self.enquiry_excel = enquiry_excel
        self.read_only_var = read_only
        if followup_excel != None and enquiry_excel != None:
            DataRead.excel_files_load = Excel_Features.load_excel_file(Followups_excel_sheet=self.followup_excel, Enquiry_excel_sheet=self.enquiry_excel, read_only=self.read_only_var)
            self.active_data = Excel_Features.active_excel_load_data(DataRead.excel_files_load)
            Excel_Features.read_excel_data(self.active_data)
        else:
            print("Parse followups excel file and enquiry excel file in DataRead Class.")
            exit()

    def __call__(self):
        Excel_Features.parse_email_data_to_thread(self.active_data)
        return DataRead.data

def main():
    data_obj = DataRead("Followup-Data1.xlsx", "Enquiry_data1.xlsx").__call__()
    print(DataRead.data)

if __name__ == '__main__':
    main()
