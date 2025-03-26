from django.core.management.base import BaseCommand
import openpyxl
from enquiries.models import (
    Followups, Enquiries, EnquiryCourses
)
from courses.models import Courses
from datetime import datetime


course_data = {
    "Amazon Web Services": {"value": 5, "name": "Amazon Web Services"},
    "Red Hat Certified Engineer": {"value": 2, "name": "Red Hat Certified Engineer"},
    "Artificial Intelligence": {"value": 21, "name": "Artificial Intelligence"},
    "Automation with Ansible": {"value": 16, "name": "Automation with Ansible"},
    "Bash shell": {"value": 22, "name": "Bash shell"},
    "Python-Basics": {"value": 4, "name": "Python-Basics"},
    "CL310 - openstack rhce": {"value": 15, "name": "CL310 - openstack rhce"},
    "Courses": {"value": 23, "name": "Courses"},
    "Devops - I": {"value": 8, "name": "Devops - I"},
    "Devops Foundation": {"value": 24, "name": "Devops Foundation"},
    "Kubernetes": {"value": 12, "name": "Kubernetes"},
    "Dot-Net": {"value": 25, "name": "Dot-Net"},
    "Redhat Virtualization": {"value": 26, "name": "Redhat Virtualization"},
    "Exam": {"value": 27, "name": ['EX200', 'EX210', "EX407"]},
    "Java-Core": {"value": 28, "name": "Java-Core"},
    "Jenkins": {"value": 8, "name": "Jenkins"},
    "Openstack - CL210": {"value": 15, "name": "Openstack - CL210"},
    "Puppet-Admin-I": {"value": 29, "name": "Puppet-Admin-I"},
    "opensthift Admin": {"value": 14, "name": "opensthift Admin"},
    "Python Advance": {"value": 30, "name": "Python Advance"},
    "Red Hat Certified System Administrator": {"value": 1, "name": "Red Hat Certified System Administrator"},
    "W3Python": {"value": 31, "name": "W3Python"}
}

location_data = {
    "AP2V Solutions": {"value": "Gurgaon", "name": "AP2V Solutions"},
    "AP2V Solutions Noida": {"value": "Noida", "name": "AP2V Solutions Noida"}
}

interest_level = {
    "Desperate to Join (Super Hot)": {"value": 1},
    "Highly Interested (Very Hot)": {"value": 2},
    "May Be Interested (Warm)": {"value": 3},
    "Interested (Hot)": {"value": 4},
    "Interested but Hesitant/Confused (Lukewarm)": {"value": 5},
    "Interest Level": {"value": 6}
}

assigned_by = {
    "Ashutosh Taiwal": {"value": 1, "name": "Ashutosh Taiwal"},
    "Assigned To": {"value": 2, "name": "Assigned To"},
    "Kirti": {"value": 3, "name": "Kirti"},
    "Manjeet": {"value": 4, "name": "Manjeet"},
    "megha": {"value": 5, "name": "megha"},
    "neha": {"value": 6, "name": "neha"},
    "Noida AP2V": {"value": 7, "name": "Noida AP2V"},
    "Pooja Kumari": {"value": 8, "name": "Pooja Kumari"},
    "Priya Saini": {"value": 9, "name": "Priya Saini"},
    "Priyanka": {"value": 10, "name": "Priyanka"},
    "Richa": {"value": 11, "name": "Richa"},
    "Vishal Saini": {"value": 12, "name": "Vishal Saini"}
}


class Command(BaseCommand):
    help = "import excel data in enquiry & followups"

    def add_arguments(self, parser):
        parser.add_argument('-e', '--enquiry', type=str, help="Enquiry Excel sheet", required=True)
        parser.add_argument('-f', '--followup', type=str, help="Followup Excel sheet", required=True)

    def handle(self, *args, **options):
        enquiry_file = options.get('enquiry')
        followup_file = options.get('followup')

        open_enquiry_excel_file = openpyxl.load_workbook(enquiry_file)
        open_followups_excel_file = openpyxl.load_workbook(followup_file)

        active_enquiry_excel_file = open_enquiry_excel_file.active
        active_followup_excel_file = open_followups_excel_file.active

        enquriy_file_max_row, enquriy_file_column_row = active_enquiry_excel_file.max_row, active_enquiry_excel_file.max_column
        followups_file_max_row, followups_file_column_row = active_followup_excel_file.max_row, active_followup_excel_file.max_column

        enquriy_file_key_values = [i for i in active_enquiry_excel_file.iter_rows(min_row=1, max_row=1, values_only=True)]
        followups_file_key_value = [i for i in active_followup_excel_file.iter_rows(min_row=1, max_row=1, values_only=True)]

        print(enquriy_file_key_values, followups_file_key_value)

        for i in range(3, enquriy_file_max_row + 1):
            enquiry_file_obj = active_enquiry_excel_file.cell(row=i, column=8)
            for p in range(3, followups_file_max_row + 1):
                followup_file_obj = active_followup_excel_file.cell(row=p, column=6)
                if enquiry_file_obj.value == followup_file_obj.value:
                    self.read_enquiry_followups_columns(
                        type=1,
                        enquiry_row_column={'row': i, 'columns': enquriy_file_column_row, 'obj': active_enquiry_excel_file},
                        followup_row_column={'row': p, 'columns': followups_file_column_row, 'obj': active_followup_excel_file}
                    )
                # else:
                #     self.read_enquiry_followups_columns(
                #         type=2,
                #         enquiry_row_column={'row': i, 'columns': enquriy_file_column_row, 'obj': active_enquiry_excel_file}
                #     )

        # print(dir(active_followup_excel_file))

    def read_enquiry_followups_columns(self, type=2, *args, **kwargs):
        '''
        type: 1 => match follups enquiry
        type: 2 => only enquiry
        '''
        if type == 1:
            enquiry_data = []
            followup_data = []
            for p in range(1, kwargs.get('enquiry_row_column')['columns'] + 1):
                enquiry_column_data = kwargs.get('enquiry_row_column')['obj'].cell(row=kwargs.get('enquiry_row_column')['row'], column=p)
                enquiry_data.append(enquiry_column_data.value)

            for i in range(1, kwargs.get('followup_row_column')['columns'] + 1):
                followup_column_data = kwargs.get('followup_row_column')['obj'].cell(row=kwargs.get('followup_row_column')['row'], column=i)
                followup_data.append(followup_column_data.value)

            print(enquiry_data, followup_data)

            self.mapping_enquiry_followups(enquiry=enquiry_data, followup=followup_data)

        else:
            enquiry_only_data = []
            for p in range(1, kwargs.get('enquiry_row_column')['columns'] + 1):
                enquiry_column_data = kwargs.get('enquiry_row_column')['obj'].cell(row=kwargs.get('enquiry_row_column')['row'], column=p)
                enquiry_only_data.append(enquiry_column_data.value)

            self.enquiry_create(enquiry=enquiry_only_data)

    def get_location(self, data):
        return location_data.get(data).get('value')

    def get_enquiry_level(self, data):
        return interest_level.get(data).get('value')

    def get_course_obj(self, obj_id):
        c = Courses.objects.filter(id=obj_id)
        if c.exists():
            return c.first()
        else:
            return Courses.objects.get(id=1)

    def get_assigned_by(self, data):
        if assigned_by.get(data):
            return assigned_by.get(data).get('value')
        return 9

    def mapping_enquiry_followups(self, *args, **kwargs):
        enquiry_list = kwargs.get('enquiry')
        followup_list = kwargs.get('followup')
        enq_obj, created_enq = Enquiries.objects.get_or_create(
            reference=followup_list[7],
            email=enquiry_list[7],
            full_name="{0} {1}".format(enquiry_list[3], enquiry_list[5]),
            mobile=enquiry_list[8],
            company_name="",
            designation="",
            enquiry_level=self.get_enquiry_level(followup_list[-3]),
            batch_days="Monday, Friday, Sunday",
            batch_time=datetime.now(),
            # assigned_by=self.get_assigned_by(followup_list[-4]),
            assigned_by_id=2,
            branch_location=self.get_location(enquiry_list[-3])
        )
        for r in followup_list[9].split(','):
            for i in course_data.keys():
                if r in [course_data[i].get("name")]:
                    enq_obj.courses.add([self.get_course_obj(course_data[i].get("value"))])
                    EnquiryCourses.objects.create(
                        enquiry_id=enq_obj,
                        courses_id=course_data[i].get("value")
                    )
        if created_enq is True:
            foll_obj, created_foll = Followups.objects.get_or_create(
                followupid=enq_obj,
                followup_mode="visit",
                response="call me response.",
                next_followup=datetime.strptime(followup_list[1], '%d-%b-%Y'),
                comments="call me response.",
                status=True,
            )
        print("mapping enquiry_id {0} followup_id {1} : created".format(enq_obj.id, foll_obj.id))

    def enquiry_create(self, *args, **kwargs):
        enquiry_list = kwargs.get('enquiry')
        course = Courses.objects.last()
        enq_obj, created_enq = Enquiries.objects.get_or_create(
            reference=enquiry_list[8],
            email=enquiry_list[5],
            full_name="{0} {1}".format(enquiry_list[2], enquiry_list[3]),
            mobile=enquiry_list[4],
            company_name="",
            designation="",
            batch_days="Monday, Friday, Sunday",
            batch_time=datetime.now(),
            branch_location=enquiry_list[7]
        )
        enq_obj.courses.set([course])
        print("Only Enquiry created id : {0}".format(enq_obj.id))
